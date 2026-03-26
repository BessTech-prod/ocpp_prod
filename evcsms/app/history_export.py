from __future__ import annotations

import io
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


DEFAULT_ORG_ID = "default"
DEFAULT_ORG_NAME = "Default"


EXPORT_COLUMNS = [
    ("transaction_id", "Transaction ID"),
    ("session_state", "Session state"),
    ("org_id", "Organization ID"),
    ("org_name", "Organization name"),
    ("charge_point", "Charge point ID"),
    ("charge_point_alias", "Charge point alias"),
    ("connectorId", "Connector"),
    ("tag", "RFID tag"),
    ("tag_alias", "RFID alias"),
    ("user_email", "User email"),
    ("user_name", "User name"),
    ("start_time", "Start time (UTC)"),
    ("stop_time", "Stop time (UTC)"),
    ("duration_minutes", "Duration (minutes)"),
    ("meter_start", "Meter start (Wh)"),
    ("meter_stop", "Meter stop (Wh)"),
    ("energy_kwh", "Energy (kWh)"),
    ("org_resolution_source", "Org resolution source"),
]


def load_json(path: Path, default):
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default


def iso_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.%fZ")


def normalize_tag(tag: str) -> str:
    return (tag or "").strip().upper()


def normalize_email(email: str | None) -> str | None:
    value = (email or "").strip().lower()
    return value or None


def normalize_org_id(org_id: str | None) -> str:
    value = (org_id or "").strip()
    return value or DEFAULT_ORG_ID


def org_display_name(org_id: str, orgs_map: Dict[str, dict]) -> str:
    org_id = normalize_org_id(org_id)
    entry = orgs_map.get(org_id) or {}
    return (entry.get("name") or org_id or DEFAULT_ORG_NAME).strip() or DEFAULT_ORG_NAME


def cp_metadata(cp_id: str, cps_map: Dict[str, Any]) -> Tuple[str, str]:
    entry = cps_map.get(cp_id)
    if isinstance(entry, dict):
        org_id = normalize_org_id(entry.get("org_id"))
        alias = (entry.get("alias") or cp_id).strip() or cp_id
        return org_id, alias
    if isinstance(entry, str):
        return normalize_org_id(entry), cp_id
    return DEFAULT_ORG_ID, cp_id


def find_user_by_email(users_map: Dict[str, dict], email: str | None) -> Tuple[str | None, dict | None]:
    wanted = normalize_email(email)
    if not wanted:
        return None, None
    for tag, user in (users_map or {}).items():
        if normalize_email(user.get("email")) == wanted:
            return normalize_tag(tag), user
    return None, None


def display_name_for_tag(tag: str, users_map: Dict[str, dict], rfids_map: Dict[str, dict]) -> str:
    tag = normalize_tag(tag)
    user = users_map.get(tag) or {}
    if not user:
        rfid = rfids_map.get(tag) or {}
        _, user = find_user_by_email(users_map, rfid.get("user_email"))
        if not user:
            return (rfid.get("alias") or tag or "Unknown").strip() or "Unknown"
    if user.get("name"):
        return str(user["name"]).strip()
    first_name = str(user.get("first_name") or "").strip()
    last_name = str(user.get("last_name") or "").strip()
    full_name = f"{first_name} {last_name}".strip()
    return full_name or tag or "Unknown"


def resolve_transaction_snapshot(
    transaction: Dict[str, Any],
    *,
    rfids_map: Dict[str, dict],
    cps_map: Dict[str, Any],
    users_map: Dict[str, dict],
    orgs_map: Dict[str, dict],
) -> Dict[str, Any]:
    tx = transaction or {}
    tag = normalize_tag(tx.get("id_tag") or tx.get("tag") or "")
    cp_id = str(tx.get("charge_point") or "").strip()
    tx_org_id = normalize_org_id(tx.get("org_id")) if tx.get("org_id") else None
    cp_org_id, cp_alias = cp_metadata(cp_id, cps_map)

    rfid = rfids_map.get(tag) or {}
    user_email = normalize_email(tx.get("user_email")) or normalize_email(rfid.get("user_email"))
    _, user = find_user_by_email(users_map, user_email)
    legacy_user = users_map.get(tag) or {}
    if not user and legacy_user:
        user = legacy_user

    if tx_org_id:
        org_id = tx_org_id
        source = "transaction"
    elif rfid.get("org_id"):
        org_id = normalize_org_id(rfid.get("org_id"))
        source = "rfid"
    elif user and user.get("org_id"):
        org_id = normalize_org_id(user.get("org_id"))
        source = "user"
    elif cp_org_id:
        org_id = normalize_org_id(cp_org_id)
        source = "charge_point"
    else:
        org_id = DEFAULT_ORG_ID
        source = "default"

    tx_org_name = (tx.get("org_name") or "").strip()
    org_name = tx_org_name or org_display_name(org_id, orgs_map)

    user_name = (tx.get("user_name") or "").strip()
    if not user_name:
        user_name = display_name_for_tag(tag, users_map, rfids_map) if tag else "Unknown"

    tag_alias = (tx.get("tag_alias") or rfid.get("alias") or tag or "Unknown").strip() or "Unknown"
    charge_point_alias = (tx.get("charge_point_alias") or cp_alias or cp_id).strip() or cp_id

    return {
        "tag": tag,
        "tag_alias": tag_alias,
        "user_email": user_email,
        "user_name": user_name,
        "org_id": org_id,
        "org_name": org_name,
        "charge_point_alias": charge_point_alias,
        "org_resolution_source": source,
    }


def enrich_transaction_snapshot(
    transaction: Dict[str, Any],
    *,
    rfids_map: Dict[str, dict],
    cps_map: Dict[str, Any],
    users_map: Dict[str, dict],
    orgs_map: Dict[str, dict],
) -> Dict[str, Any]:
    enriched = dict(transaction or {})
    enriched.update(
        resolve_transaction_snapshot(
            enriched,
            rfids_map=rfids_map,
            cps_map=cps_map,
            users_map=users_map,
            orgs_map=orgs_map,
        )
    )
    return enriched


def parse_iso8601(value: Any) -> datetime | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except Exception:
        return None


def session_state(transaction: Dict[str, Any]) -> str:
    if transaction.get("stop_time") and transaction.get("meter_stop") is not None:
        return "completed"
    return "open"


def compute_energy_kwh(transaction: Dict[str, Any]) -> float | None:
    try:
        meter_start = float(transaction.get("meter_start"))
        meter_stop = float(transaction.get("meter_stop"))
    except Exception:
        return None
    return round(max(0.0, meter_stop - meter_start) / 1000.0, 3)


def compute_duration_minutes(transaction: Dict[str, Any]) -> float | None:
    start_dt = parse_iso8601(transaction.get("start_time"))
    stop_dt = parse_iso8601(transaction.get("stop_time"))
    if not start_dt or not stop_dt:
        return None
    seconds = (stop_dt - start_dt).total_seconds()
    return round(max(0.0, seconds) / 60.0, 1)


def load_backup_source(data_dir: Path, config_dir: Path) -> Dict[str, Any]:
    return {
        "transactions": load_json(data_dir / "transactions.json", []),
        "orgs": load_json(config_dir / "orgs.json", {}),
        "rfids": load_json(config_dir / "rfids.json", {}),
        "users": load_json(config_dir / "users.json", {}),
        "cps": load_json(config_dir / "cps.json", {}),
    }


def build_backup_rows(source: Dict[str, Any]) -> List[Dict[str, Any]]:
    txs = source.get("transactions") or []
    rfids = source.get("rfids") or {}
    users = source.get("users") or {}
    orgs = source.get("orgs") or {}
    cps = source.get("cps") or {}

    rows: List[Dict[str, Any]] = []
    for tx in txs:
        enriched = enrich_transaction_snapshot(
            tx,
            rfids_map=rfids,
            cps_map=cps,
            users_map=users,
            orgs_map=orgs,
        )
        energy_kwh = compute_energy_kwh(enriched)
        row = {
            "transaction_id": enriched.get("transaction_id"),
            "session_state": session_state(enriched),
            "org_id": enriched.get("org_id"),
            "org_name": enriched.get("org_name"),
            "charge_point": enriched.get("charge_point"),
            "charge_point_alias": enriched.get("charge_point_alias"),
            "connectorId": enriched.get("connectorId"),
            "tag": enriched.get("tag") or normalize_tag(enriched.get("id_tag") or ""),
            "tag_alias": enriched.get("tag_alias"),
            "user_email": enriched.get("user_email"),
            "user_name": enriched.get("user_name"),
            "start_time": enriched.get("start_time"),
            "stop_time": enriched.get("stop_time"),
            "duration_minutes": compute_duration_minutes(enriched),
            "meter_start": enriched.get("meter_start"),
            "meter_stop": enriched.get("meter_stop"),
            "energy_kwh": energy_kwh,
            "org_resolution_source": enriched.get("org_resolution_source"),
        }
        rows.append(row)

    rows.sort(
        key=lambda row: (
            row.get("org_name") or "",
            row.get("stop_time") or row.get("start_time") or "",
            str(row.get("transaction_id") or ""),
        ),
        reverse=False,
    )
    return rows


def build_backup_manifest(rows: List[Dict[str, Any]], generated_at: str) -> Dict[str, Any]:
    summary: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        org_id = row.get("org_id") or DEFAULT_ORG_ID
        bucket = summary.setdefault(
            org_id,
            {
                "org_id": org_id,
                "org_name": row.get("org_name") or org_id,
                "sessions_total": 0,
                "sessions_completed": 0,
                "sessions_open": 0,
                "energy_kwh_completed": 0.0,
                "first_activity": None,
                "latest_activity": None,
            },
        )
        bucket["sessions_total"] += 1
        if row.get("session_state") == "completed":
            bucket["sessions_completed"] += 1
        else:
            bucket["sessions_open"] += 1
        if row.get("energy_kwh") is not None:
            bucket["energy_kwh_completed"] = round(bucket["energy_kwh_completed"] + float(row["energy_kwh"]), 3)

        for field_name in ("start_time", "stop_time"):
            value = row.get(field_name)
            if not value:
                continue
            if not bucket["first_activity"] or value < bucket["first_activity"]:
                bucket["first_activity"] = value
            if not bucket["latest_activity"] or value > bucket["latest_activity"]:
                bucket["latest_activity"] = value

    return {
        "generated_at": generated_at,
        "organizations": sorted(summary.values(), key=lambda item: item["org_name"] or item["org_id"]),
        "totals": {
            "organizations": len(summary),
            "sessions_total": len(rows),
            "sessions_completed": sum(1 for row in rows if row.get("session_state") == "completed"),
            "sessions_open": sum(1 for row in rows if row.get("session_state") != "completed"),
            "energy_kwh_completed": round(sum(float(row.get("energy_kwh") or 0.0) for row in rows), 3),
        },
    }


def _safe_sheet_title(name: str, used_titles: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", " ", str(name or "Unknown")).strip()
    base = re.sub(r"\s+", " ", base) or "Unknown"
    title = base[:31]
    if title not in used_titles:
        used_titles.add(title)
        return title

    counter = 2
    while True:
        suffix = f" ({counter})"
        trimmed = base[: max(1, 31 - len(suffix))].rstrip()
        candidate = f"{trimmed}{suffix}"
        if candidate not in used_titles:
            used_titles.add(candidate)
            return candidate
        counter += 1


def _autosize_worksheet(ws):
    for idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 40)


def build_history_workbook(rows: List[Dict[str, Any]], generated_at: str | None = None) -> bytes:
    generated_at = generated_at or iso_now()
    manifest = build_backup_manifest(rows, generated_at)

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws.freeze_panes = "A2"
    summary_ws.append([
        "Generated at (UTC)",
        "Organization ID",
        "Organization name",
        "Total sessions",
        "Completed sessions",
        "Open sessions",
        "Completed energy (kWh)",
        "First activity",
        "Latest activity",
    ])
    for item in manifest["organizations"]:
        summary_ws.append([
            generated_at,
            item["org_id"],
            item["org_name"],
            item["sessions_total"],
            item["sessions_completed"],
            item["sessions_open"],
            item["energy_kwh_completed"],
            item["first_activity"],
            item["latest_activity"],
        ])
    _autosize_worksheet(summary_ws)
    summary_ws.auto_filter.ref = summary_ws.dimensions

    by_org: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        by_org.setdefault(row.get("org_id") or DEFAULT_ORG_ID, []).append(row)

    used_titles = {summary_ws.title}
    if not by_org:
        empty_ws = wb.create_sheet("No data")
        empty_ws.append(["No charging transactions available for backup."])
        _autosize_worksheet(empty_ws)
    else:
        for org_id, org_rows in sorted(by_org.items(), key=lambda item: (item[1][0].get("org_name") or item[0], item[0])):
            org_name = org_rows[0].get("org_name") or org_id
            ws = wb.create_sheet(_safe_sheet_title(org_name, used_titles))
            ws.freeze_panes = "A2"
            ws.append([label for _, label in EXPORT_COLUMNS])
            for row in org_rows:
                ws.append([row.get(key) for key, _ in EXPORT_COLUMNS])
            _autosize_worksheet(ws)
            ws.auto_filter.ref = ws.dimensions

    metadata_ws = wb.create_sheet("Metadata")
    metadata_ws.append(["Generated at (UTC)", generated_at])
    metadata_ws.append(["Organizations", manifest["totals"]["organizations"]])
    metadata_ws.append(["Sessions total", manifest["totals"]["sessions_total"]])
    metadata_ws.append(["Sessions completed", manifest["totals"]["sessions_completed"]])
    metadata_ws.append(["Sessions open", manifest["totals"]["sessions_open"]])
    metadata_ws.append(["Completed energy (kWh)", manifest["totals"]["energy_kwh_completed"]])
    _autosize_worksheet(metadata_ws)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

